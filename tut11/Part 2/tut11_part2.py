from openpyxl import load_workbook


# Load the first Excel file with two sheets
file1 = 'Output-1.xlsx'


# Load the second Excel file with one sheet
file2 = 'demo1.xlsx'

wb1 = load_workbook(file1)
wb2 = load_workbook(file2)

# Create a new workbook to combine sheets
wb_combined = load_workbook(file1)  # Start by using the first file's content

# Copy sheets from the second file
for sheet_name in wb2.sheetnames:
    sheet = wb2[sheet_name]
    new_sheet = wb_combined.create_sheet(sheet_name)
    
    # Copy all the content from the sheet (including values, styles, etc.)
    for row in sheet.rows:
        for cell in row:
            new_sheet[cell.coordinate].value = cell.value
            if cell.has_style:
                new_sheet[cell.coordinate].style = cell.style

    # Copy images and other non-data objects (images/charts)
    for image in sheet._images:
        new_sheet.add_image(image)

# Save the combined workbook as a new file
wb_combined.save('Output 2 Lab 11.xlsx')
