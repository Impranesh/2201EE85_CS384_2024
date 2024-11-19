
import ast

inputfile = 'C:\Users\PAVILION\Downloads/2201EE85_CS384_2024/tut09/python dates.txt'
Classes_Taken = []
Classes_Not_Taken = []
Exams = []
Class_Timing="18:00 - 20:00"
with open(inputfile, 'r') as f:
    for line in f:
        if "classes_taken_dates" in line:
            Classes_Taken = ast.literal_eval(line.split('=')[1].strip())
        elif "classes_missed_dates" in line:
            Classes_Not_Taken = ast.literal_eval(line.split('=')[1].strip())
        elif "exams_dates" in line:
            Exams = ast.literal_eval(line.split('=')[1].strip())

print("Classes Taken:", Classes_Taken)
print("Classes Not Taken:", Classes_Not_Taken)
print("Exams:", Exams)
print("Class_Timing:",Class_Timing)


for i in Classes_Not_Taken:
  Classes_Taken.append(i)
for i in Exams:
  Classes_Taken.append(i)

Classes_Taken = sorted(Classes_Taken)



inputfile='C:/Users/PAVILION/Downloads/2201EE85_CS384_2024/tut09/stud_list.txt'
dict_list={}
with open(inputfile, 'r') as f:
    for line in f:
      a=line.split(' ',1)
      dict_list[a[0]]=a[1].strip()
print(dict_list)




inputfile='C:/Users/PAVILION/Downloads/2201EE85_CS384_2024/tut09/input_attendance.csv'
import pandas as pd
a=[]

with open(inputfile, 'r') as f:
  for line in f:
    b=line.split(',')
    b[1]=b[1].strip()
    a.append(b)
entry_to_remove=['06/08/2024 18:24:11','2201EE85 Pranesh Priyanshu']
if entry_to_remove in a:
  a.remove(entry_to_remove)

print(a)



def validate_roll(roll):
    """
    Validates the roll number using a regular expression.
    Adjust the regex pattern based on your roll number format.
    """
    # Example pattern: starts with 4 digits, followed by 2 uppercase letters, ending with 2-3 digits
    pattern = r'^\d{4}[A-Z]{2}\d{2,3}$'
    return re.match(pattern, roll) is not None


def extract_roll_no(roll_entry):
    """
    Extracts the roll number from the roll entry, which may include the name.
    Assumes the roll number is the first word.
    """
    if not isinstance(roll_entry, str):
        return None
    parts = roll_entry.strip().split(' ', 1)
    roll_no = parts[0] if len(parts) > 0 else None
    return roll_no if validate_roll(roll_no) else None





import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl import Workbook, load_workbook
from datetime import datetime, time
import re



output_excel_file = 'C:/Users/PAVILION/Downloads/2201EE85_CS384_2024/tut09/output_excel.xlsx'  # Update the path as needed





def process_attendance(attendance_data, student_dict, classes_taken_dates):

    # Initialize attendance_record
    attendance_record = {
        roll: {
            'Name': student_dict.get(roll, ""),
            'Attendance': {date: 0 for date in classes_taken_dates},
            'Proxy': 0,
            'Proxy Dates':[],
            'Total Classes': 2*len(classes_taken_dates)-4,
            'Total Classes Attended': 0
        } for roll in student_dict.keys()
    }

    # Iterate through each attendance entry
    for entry in attendance_data:
        timestamp_str, roll_entry = entry

        roll_no = extract_roll_no(roll_entry)

        if not roll_no:
            # Invalid roll number, skip this entry
            print(f"Invalid or missing roll number in entry: {entry}")
            continue

        # Parse the timestamp

        try:
            timestamp = datetime.strptime(timestamp_str, '%d/%m/%Y %H:%M:%S')
        except ValueError:
            # Invalid timestamp format, skip this entry
            print(f"Invalid timestamp format in entry: {entry}")
            continue

        # Check if the date is a class date

        date_str = timestamp.strftime('%d/%m/%Y')
        if date_str not in classes_taken_dates:
            attendance_record[roll_no]['Proxy'] += 1
            attendance_record[roll_no]['Proxy Dates'].append(date_str)
            # Not a class date, skip
            continue

        # Check if the time is between 6:00 PM and 8:00 PM
        if time(18, 0, 0) <= timestamp.time() <= time(20, 0, 0):
            # Within attendance window
            if roll_no in attendance_record:
                current_count = attendance_record[roll_no]['Attendance'][date_str]
                if current_count < 2:
                    attendance_record[roll_no]['Attendance'][date_str] += 1
                    # Increment Total Classes Attended only once per date
                else:
                    # Exceeds allowed attendance count
                    attendance_record[roll_no]['Attendance'][date_str] =3
                    attendance_record[roll_no]['Proxy'] += 1
                    attendance_record[roll_no]['Proxy Dates'].append(date_str)
            else:
                # Roll number not in student_dict, possibly due to add/drop, ignore
                print(f"Roll number {roll_no} not found in student dictionary.")
                continue
        else:
            # Outside attendance window, count as proxy
            if roll_no in attendance_record:
                attendance_record[roll_no]['Proxy'] += 1
                attendance_record[roll_no]['Proxy Dates'].append(date_str)
            else:
                # Roll number not in student_dict, possibly due to add/drop, ignore
                print(f"Roll number {roll_no} not found in student dictionary.")
                continue

    # After processing all entries, calculate Total Classes Attended
    for roll, info in attendance_record.items():
         for count in info['Attendance'].values():
            attendance_record[roll]['Total Classes Attended'] +=count

    return attendance_record






attendance_record = process_attendance(a, dict_list, Classes_Taken)
print("Attendance processing completed.")



def generate_excel(attendance_record, classes_taken_dates, output_file):
    """
    Generates an Excel file from the attendance_record with conditional formatting.
    Includes 'Proxy Dates', 'Total Classes', and 'Total Classes Attended' columns.
    """
    # Create a list of dictionaries for DataFrame
    data = []
    for roll, info in attendance_record.items():
        row = {
            'Roll Number': roll,
            'Name': info['Name']
        }
        for date in classes_taken_dates:
            row[date] = info['Attendance'][date]
        row['Proxy'] = info['Proxy']
        # Convert Proxy Dates set to comma-separated string
        row['Proxy Dates'] = ', '.join(sorted(info['Proxy Dates']))
        row['Total Classes'] = info['Total Classes']
        row['Total Classes Attended'] = info['Total Classes Attended']
        row['Percentage'] = (info['Total Classes Attended'] / info['Total Classes']) * 100
        data.append(row)

    # Create DataFrame
    df_output = pd.DataFrame(data)

    # Write DataFrame to Excel
    try:
        df_output.to_excel(output_file, index=False)
        print(f"Excel file '{output_file}' created successfully.")
    except Exception as e:
        print(f"Error writing to Excel file: {e}")
        return

    # Load workbook for formatting
    try:
        wb = load_workbook(output_file)
        ws = wb.active
    except Exception as e:
        print(f"Error loading Excel file for formatting: {e}")
        return

    # Define fill colors
    fill_absent = PatternFill(start_color="FAA0A0", end_color="FAA0A0", fill_type="solid")    # Red
    fill_partial = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")   # Yellow
    fill_full = PatternFill(start_color="97DB4F", end_color="97DB4F", fill_type="solid")      # Green
    fill_proxy = PatternFill(start_color="13b5e0", end_color="ADD8E6", fill_type="solid")    # Light Blue

    # Apply conditional formatting to attendance columns
    # Assuming 'Roll Number' is column 1, 'Name' is column 2
    # Attendance dates start from column 3, Proxy is the last column before 'Proxy Dates'
    attendance_start_col = 3
    attendance_end_col = 2 + len(classes_taken_dates)
    proxy_col = attendance_end_col + 1
    proxy_dates_col = proxy_col + 1
    total_classes_col = proxy_dates_col + 1
    total_attended_col = total_classes_col + 1
    percentage_col = total_attended_col + 1

    # Apply formatting to attendance counts
    for row in ws.iter_rows(min_row=2, min_col=attendance_start_col, max_col=attendance_end_col, max_row=ws.max_row):
        for cell in row:
            if cell.value == 0:
                cell.fill = fill_absent
            elif cell.value == 1:
                cell.fill = fill_partial
            elif cell.value == 2:
                cell.fill = fill_full
            elif cell.value==3:
                cell.value=2
                cell.fill=fill_proxy
            # Else, no fill

    # Apply formatting to Proxy column
    for row in ws.iter_rows(min_row=2, min_col=proxy_col, max_col=proxy_col, max_row=ws.max_row):
        for cell in row:
            if cell.value > 0:
                cell.fill = fill_proxy  # Light Blue for proxies

   # Apply formatting to Percentage column
   # If it is less than 75 % then mark it as red
    for row in ws.iter_rows(min_row=2, min_col=percentage_col, max_col=percentage_col, max_row=ws.max_row):
        for cell in row:
            if cell.value <75.0:
              cell.fill = fill_absent

    # Save the formatted workbook
    try:
        wb.save(output_file)
        print(f"Excel file '{output_file}' formatted with conditional highlights successfully.")
    except Exception as e:
        print(f"Error saving formatted Excel file: {e}")




generate_excel(attendance_record, Classes_Taken, output_excel_file)
print("Excel file generation completed.")